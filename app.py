"""
=============================================================================
GAMBLING OSINT — Web Interface
=============================================================================
Simple Flask app with two steps:
  Step 1: Upload CSV/Excel → Get NEW main gambling domains (deduplicated)
  Step 2: Upload that CSV  → Get infrastructure info (like main2.py) for each

Run:  python app.py
Open: http://localhost:5000
=============================================================================
"""

import os
import io
import csv
import time
import uuid
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

from flask import Flask, request, jsonify, render_template, send_file

import dns.resolver
import tldextract
from ipwhois import IPWhois

import config
from utils import normalize_domain, is_valid_domain, read_excel_domains

# Import expansion logic from quick_expand
from quick_expand import (
    generate_candidates,
    check_dns,
    GAMBLING_KEYWORDS,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

# Store generated files temporarily
TEMP_DIR = os.path.join(config.OUTPUT_DIR, "temp")
os.makedirs(TEMP_DIR, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def read_domains_from_upload(file_storage) -> set:
    """
    Read domains from an uploaded file (Excel, CSV, or TXT).
    Smart detection: scans ALL columns for anything that looks like a domain/URL.
    Works with any Excel/CSV format, not just the specific 'URL of sites' column.
    """
    filename = file_storage.filename.lower()
    domains = set()

    if filename.endswith((".xlsx", ".xls")):
        tmp_path = os.path.join(TEMP_DIR, f"upload_{uuid.uuid4().hex}.xlsx")
        file_storage.save(tmp_path)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Skip Chartsheets
                if not hasattr(ws, 'iter_rows'):
                    continue

                for row in ws.iter_rows(min_row=1):
                    for cell in row:
                        val = str(cell.value).strip() if cell.value else ""
                        if not val or val.lower() in ("none", ""):
                            continue
                        # Try to extract a domain from every cell value
                        d = normalize_domain(val)
                        if d and is_valid_domain(d):
                            domains.add(d)

            wb.close()
        except ImportError:
            logger.error(
                "openpyxl is not installed! "
                "Run: pip install -r requirements.txt"
            )
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    elif filename.endswith(".csv"):
        text = file_storage.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            # Scan ALL columns for domain-like values
            for key, val in row.items():
                if val:
                    d = normalize_domain(val.strip())
                    if d and is_valid_domain(d):
                        domains.add(d)

    else:
        # Plain text — one domain per line
        text = file_storage.read().decode("utf-8")
        for line in text.strip().split("\n"):
            line = line.strip()
            if line and not line.startswith("#"):
                d = normalize_domain(line)
                if d and is_valid_domain(d):
                    domains.add(d)

    logger.info(f"Extracted {len(domains)} domains from uploaded file: {filename}")
    return domains


def deduplicate_to_main_domains(domain_list: list) -> list:
    """
    Collapse domain variations to one per brand.
    e.g. betway.com, betway.in, betway.net → keep betway.com
    """
    TLD_PRIORITY = {
        "com": 1, "in": 2, "co.in": 3, "net": 4, "org": 5,
        "live": 6, "online": 7, "bet": 8, "io": 9, "co": 10,
    }

    brands = {}  # brand_name → {domain, ip, priority}

    for entry in domain_list:
        domain = entry["domain"]
        ext = tldextract.extract(domain)
        brand = ext.domain.lower()
        tld = ext.suffix.lower()
        priority = TLD_PRIORITY.get(tld, 50)

        if brand not in brands or priority < brands[brand]["priority"]:
            brands[brand] = {
                "domain": domain,
                "ip_address": entry.get("ip_address", ""),
                "priority": priority,
            }

    result = []
    for brand, info in sorted(brands.items()):
        result.append({
            "domain": info["domain"],
            "ip_address": info["ip_address"],
        })

    return result


def lookup_single_domain(domain: str) -> dict:
    """Run main2.py-style infrastructure lookup for one domain."""
    ext = tldextract.extract(domain)
    main_domain = f"{ext.domain}.{ext.suffix}"

    result = {
        "domain": domain,
        "main_domain": main_domain,
        "ipv4_addresses": "",
        "ipv6_addresses": "",
        "hosting_provider": "",
        "country": "",
        "asn": "",
        "asn_description": "",
    }

    resolver = dns.resolver.Resolver()
    resolver.nameservers = ["8.8.8.8", "1.1.1.1"]
    resolver.timeout = 5
    resolver.lifetime = 8

    # IPv4
    ipv4_list = []
    try:
        answers = resolver.resolve(main_domain, "A", lifetime=8)
        ipv4_list = [a.to_text() for a in answers]
        result["ipv4_addresses"] = "; ".join(ipv4_list)
    except Exception:
        pass

    # IPv6
    try:
        answers = resolver.resolve(main_domain, "AAAA", lifetime=8)
        result["ipv6_addresses"] = "; ".join(a.to_text() for a in answers)
    except Exception:
        pass

    # RDAP/Whois
    if ipv4_list:
        try:
            obj = IPWhois(ipv4_list[0])
            rdap = obj.lookup_rdap()
            network = rdap.get("network", {})
            result["hosting_provider"] = network.get("name") or ""
            result["country"] = network.get("country") or ""
            result["asn"] = rdap.get("asn") or ""
            result["asn_description"] = rdap.get("asn_description") or ""
        except Exception:
            pass

    return result


# ═══════════════════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/expand", methods=["POST"])
def api_expand():
    """Step 1: Upload file → expand → return deduplicated main domains."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    uploaded = request.files["file"]
    if not uploaded.filename:
        return jsonify({"error": "Empty filename"}), 400

    start = time.time()

    # 1. Read seeds
    seeds = read_domains_from_upload(uploaded)
    if not seeds:
        return jsonify({"error": "No valid domains found in file"}), 400

    # 2. Generate candidates
    candidates = generate_candidates(seeds)

    # 3. Parallel DNS validation
    live = []
    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = {executor.submit(check_dns, d): d for d in candidates}
        for future in as_completed(futures):
            domain, ip = future.result()
            if ip:
                live.append({"domain": domain, "ip_address": ip})

    # 4. Deduplicate to main domains
    main_domains = deduplicate_to_main_domains(live)

    elapsed = round(time.time() - start, 1)

    # 5. Save CSV for download
    csv_id = uuid.uuid4().hex[:8]
    csv_path = os.path.join(TEMP_DIR, f"expanded_{csv_id}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["domain", "ip_address"])
        writer.writeheader()
        writer.writerows(main_domains)

    return jsonify({
        "success": True,
        "seeds": len(seeds),
        "candidates_checked": len(candidates),
        "total_resolved": len(live),
        "main_domains": len(main_domains),
        "elapsed_seconds": elapsed,
        "download_id": csv_id,
        "domains": main_domains,
    })


@app.route("/api/lookup", methods=["POST"])
def api_lookup():
    """Step 2: Upload domains CSV → infrastructure lookup for each."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    uploaded = request.files["file"]
    if not uploaded.filename:
        return jsonify({"error": "Empty filename"}), 400

    start = time.time()

    # 1. Read domains
    domains = read_domains_from_upload(uploaded)
    if not domains:
        return jsonify({"error": "No valid domains found in file"}), 400

    # 2. Lookup each (threaded for speed, but limited to avoid RDAP blocks)
    results = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(lookup_single_domain, d): d for d in sorted(domains)}
        for future in as_completed(futures):
            results.append(future.result())

    results.sort(key=lambda x: x["domain"])
    elapsed = round(time.time() - start, 1)

    # 3. Save CSV
    csv_id = uuid.uuid4().hex[:8]
    csv_path = os.path.join(TEMP_DIR, f"lookup_{csv_id}.csv")
    fieldnames = [
        "domain", "main_domain", "ipv4_addresses", "ipv6_addresses",
        "hosting_provider", "country", "asn", "asn_description",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    return jsonify({
        "success": True,
        "total_domains": len(domains),
        "elapsed_seconds": elapsed,
        "download_id": csv_id,
        "results": results,
    })


@app.route("/api/download/<download_id>")
def api_download(download_id):
    """Download a generated CSV file."""
    # Check both prefixes
    for prefix in ["expanded_", "lookup_"]:
        path = os.path.join(TEMP_DIR, f"{prefix}{download_id}.csv")
        if os.path.exists(path):
            return send_file(
                path,
                mimetype="text/csv",
                as_attachment=True,
                download_name=f"{prefix}{download_id}.csv",
            )
    return jsonify({"error": "File not found"}), 404


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Check dependencies first
    missing = []
    for pkg in ["flask", "dns.resolver", "tldextract", "openpyxl", "ipwhois"]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg.replace("dns.resolver", "dnspython"))

    if missing:
        print("\n  ERROR: Missing dependencies!")
        print(f"  Missing: {', '.join(missing)}")
        print("  Fix: pip install -r requirements.txt\n")
        exit(1)

    print()
    print("=" * 50)
    print("  GAMBLING OSINT - Web Interface")
    print("  Open: http://localhost:5000")
    print("=" * 50)
    print()
    app.run(debug=True, host="0.0.0.0", port=5000)
