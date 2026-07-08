"""
=============================================================================
GAMBLING OSINT - UTILITY MODULE
=============================================================================
Shared utility functions used across all pipeline stages:
- Domain normalization and validation
- CSV/Excel I/O helpers
- Deduplication logic
- Progress tracking
=============================================================================
"""

import re
import csv
import logging
import tldextract
from urllib.parse import urlparse

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# DOMAIN NORMALIZATION & VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

def normalize_domain(raw_input: str) -> str:
    """
    Normalize a raw domain/URL string to its registrable domain.
    
    Examples:
        'https://www.example.com/path?q=1' → 'example.com'
        'http://sub.domain.co.in/page'     → 'domain.co.in'
        'EXAMPLE.COM'                      → 'example.com'
    
    Returns empty string if extraction fails.
    """
    if not raw_input or not isinstance(raw_input, str):
        return ""

    raw_input = raw_input.strip()

    # Remove protocol prefix
    cleaned = raw_input.replace("https://", "").replace("http://", "")
    # Remove path and query
    cleaned = cleaned.split("/")[0].split("?")[0].split("#")[0]
    # Remove port
    cleaned = cleaned.split(":")[0]
    # Lowercase
    cleaned = cleaned.lower().strip()

    # Use tldextract for robust extraction
    try:
        ext = tldextract.extract(cleaned)
        if ext.domain and ext.suffix:
            return f"{ext.domain}.{ext.suffix}"
        elif ext.domain:
            return ext.domain
    except Exception as e:
        logger.debug(f"tldextract failed for '{raw_input}': {e}")

    return cleaned if cleaned else ""


def extract_full_domain(raw_input: str) -> str:
    """
    Extract the full domain including subdomain (but no path/query).
    
    Example: 'https://www.sub.example.com/path' → 'www.sub.example.com'
    """
    if not raw_input or not isinstance(raw_input, str):
        return ""

    raw_input = raw_input.strip()
    cleaned = raw_input.replace("https://", "").replace("http://", "")
    cleaned = cleaned.split("/")[0].split("?")[0].split("#")[0]
    cleaned = cleaned.split(":")[0]

    return cleaned.lower().strip()


def is_valid_domain(domain: str) -> bool:
    """
    Check whether a string looks like a valid domain name.
    Basic validation: contains a dot, no spaces, valid characters,
    and filters out plain IP addresses and IP-like domains (e.g., 154.198.173.1.co).
    """
    if not domain or not isinstance(domain, str):
        return False

    domain = domain.strip().lower()

    # Must contain at least one dot
    if "." not in domain:
        return False

    # No spaces allowed
    if " " in domain:
        return False

    # Basic pattern: alphanumeric, hyphens, dots
    pattern = r'^[a-z0-9]([a-z0-9\-]*[a-z0-9])?(\.[a-z0-9]([a-z0-9\-]*[a-z0-9])?)+$'
    if not bool(re.match(pattern, domain)):
        return False

    # Filter out plain IPv4 addresses (e.g. 192.168.1.1)
    ip_pattern = r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
    if re.match(ip_pattern, domain):
        return False

    # Filter out IP-like domains (e.g. 154.198.173.1.co)
    try:
        ext = tldextract.extract(domain)
        non_tld_parts = []
        if ext.subdomain:
            non_tld_parts.extend(ext.subdomain.split('.'))
        if ext.domain:
            non_tld_parts.append(ext.domain)

        # If we have 2 or more non-TLD segments and all are purely numeric, it's an IP structure
        if len(non_tld_parts) >= 2 and all(part.isdigit() for part in non_tld_parts):
            return False
    except Exception:
        pass

    return True


def deduplicate_domains(domain_list: list) -> list:
    """
    Remove duplicate domains while preserving order.
    Normalizes before comparing.
    """
    seen = set()
    unique = []
    for d in domain_list:
        normalized = normalize_domain(d)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique.append(normalized)
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# CSV HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def read_csv_domains(filepath: str, domain_column: str = "domain") -> list:
    """
    Read domains from a CSV file. Returns list of dicts.
    """
    results = []
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                results.append(dict(row))
        logger.info(f"Read {len(results)} rows from {filepath}")
    except FileNotFoundError:
        logger.warning(f"CSV file not found: {filepath}")
    except Exception as e:
        logger.error(f"Error reading CSV {filepath}: {e}")
    return results


def write_csv(filepath: str, data: list, fieldnames: list = None):
    """
    Write a list of dicts to a CSV file.
    If fieldnames is None, uses keys from the first dict.
    """
    if not data:
        logger.warning(f"No data to write to {filepath}")
        return

    if fieldnames is None:
        # Collect all keys across all dicts to handle inconsistent keys
        all_keys = []
        seen_keys = set()
        for row in data:
            for key in row.keys():
                if key not in seen_keys:
                    all_keys.append(key)
                    seen_keys.add(key)
        fieldnames = all_keys

    try:
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        logger.info(f"Wrote {len(data)} rows to {filepath}")
    except Exception as e:
        logger.error(f"Error writing CSV {filepath}: {e}")


def append_csv(filepath: str, rows: list, fieldnames: list):
    """
    Append rows to an existing CSV. Creates the file with headers if it doesn't exist.
    """
    import os
    file_exists = os.path.exists(filepath)

    try:
        with open(filepath, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            if not file_exists:
                writer.writeheader()
            writer.writerows(rows)
    except Exception as e:
        logger.error(f"Error appending to CSV {filepath}: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def read_excel_domains(filepath: str) -> list:
    """
    Read domains from the existing Excel spreadsheet format.
    Expected columns: 'Betting Site names', 'URL of sites', 'IPv4 Address',
                      'IPv6 Address', 'Host name'
    Returns list of dicts with normalized keys.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        results = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip Chartsheets and other non-data sheets
            if not hasattr(ws, 'iter_rows'):
                logger.debug(f"Skipping non-data sheet: {sheet_name}")
                continue

            # Read headers from first row
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = str(cell.value).strip() if cell.value else ""
                headers.append(val)

            if not headers or not any(headers):
                continue

            for row in ws.iter_rows(min_row=2):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = (
                            str(cell.value).strip() if cell.value else ""
                        )

                # Extract URL
                url = row_data.get("URL of sites", "")
                if not url:
                    continue

                domain = normalize_domain(url)
                if domain:
                    results.append({
                        "site_name": row_data.get("Betting Site names ", "").strip(),
                        "url": url.strip(),
                        "domain": domain,
                        "ipv4_existing": row_data.get("IPv4 Address", ""),
                        "ipv6_existing": row_data.get("IPv6 Address", ""),
                        "host_existing": row_data.get("Host name ", ""),
                    })

        wb.close()
        logger.info(f"Read {len(results)} domains from Excel: {filepath}")
        return results

    except Exception as e:
        logger.error(f"Error reading Excel {filepath}: {e}")
        return []


# ─────────────────────────────────────────────────────────────────────────────
# PROGRESS TRACKING
# ─────────────────────────────────────────────────────────────────────────────

class ProgressTracker:
    """Simple progress tracker with logging."""

    def __init__(self, total: int, task_name: str = "Processing"):
        self.total = total
        self.task_name = task_name
        self.current = 0
        self.successes = 0
        self.failures = 0
        self.skipped = 0

    def update(self, success: bool = True, skipped: bool = False):
        self.current += 1
        if skipped:
            self.skipped += 1
        elif success:
            self.successes += 1
        else:
            self.failures += 1

        # Log progress every 10% or every 10 items for small sets
        interval = max(1, self.total // 10)
        if self.current % interval == 0 or self.current == self.total:
            pct = (self.current / self.total) * 100 if self.total > 0 else 0
            logger.info(
                f"{self.task_name}: {self.current}/{self.total} ({pct:.0f}%) "
                f"[OK={self.successes} FAIL={self.failures} SKIP={self.skipped}]"
            )

    def summary(self) -> dict:
        return {
            "total": self.total,
            "processed": self.current,
            "successes": self.successes,
            "failures": self.failures,
            "skipped": self.skipped,
        }
